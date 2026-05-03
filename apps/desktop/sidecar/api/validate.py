from fastapi import APIRouter
from pydantic import BaseModel
from pathlib import Path
from zipfile import ZipFile
from typing import Dict, List, Optional

router = APIRouter()


class ValidateTemplateRequest(BaseModel):
    template_path: str


class ValidateExcelRequest(BaseModel):
    input_path: str
    config: dict


class ValidateConfigRequest(BaseModel):
    config: dict


class CrossValidateRequest(BaseModel):
    template_path: str
    excel_path: str
    config: dict


def _issue(severity, source, code, message, **kwargs):
    d = {"severity": severity, "source": source, "code": code, "message": message}
    loc = {}
    for k in ("sheet", "column", "field", "placeholder", "flag"):
        if k in kwargs:
            loc[k] = kwargs[k]
    if loc:
        d["location"] = loc
    if "suggestion" in kwargs:
        d["suggestion"] = kwargs["suggestion"]
    return d


# ── POST /api/validate-template ──

@router.post("/validate-template")
def validate_template(req: ValidateTemplateRequest):
    issues = []
    path = Path(req.template_path)

    if not path.exists():
        return {"valid": False, "issues": [_issue("error", "template", "FILE_NOT_FOUND", f"模板文件不存在: {req.template_path}")], "warnings": []}

    if path.suffix.lower() != ".docx":
        issues.append(_issue("warning", "template", "WRONG_EXTENSION", f"文件扩展名为 .{path.suffix.lstrip('.')}，期望 .docx"))

    try:
        with ZipFile(path) as zf:
            names = zf.namelist()
            if "word/document.xml" not in names:
                issues.append(_issue("error", "template", "INVALID_DOCX", "不是有效的 Word 文档：缺少 word/document.xml"))
    except Exception as e:
        return {"valid": False, "issues": [_issue("error", "template", "UNREADABLE", f"无法读取文件（不是有效的 ZIP）: {e}")], "warnings": []}

    # Parse template structure
    try:
        from report_engine.template_parser import parse_template
        structure, parse_warnings = parse_template(str(path))
        for w in parse_warnings:
            issues.append(_issue("warning", "template", "PARSE_WARNING", w))
    except Exception as e:
        issues.append(_issue("error", "template", "PARSE_ERROR", f"模板解析失败: {e}"))
        structure = None

    # Check styles
    try:
        from report_engine.style_checker import check_template_styles
        style_result = check_template_styles(str(path))
        for s in style_result.missing:
            issues.append(_issue("warning", "template", "MISSING_STYLE", f"模板缺少样式: {s}", suggestion=f"请在 Word 中添加样式 '{s}'"))
        for s in style_result.wrong_type:
            issues.append(_issue("warning", "template", "WRONG_STYLE_TYPE", f"样式类型不正确: {s}"))
    except Exception:
        pass

    has_errors = any(i["severity"] == "error" for i in issues)
    return {
        "valid": not has_errors,
        "structure": structure,
        "issues": issues,
        "warnings": [i["message"] for i in issues if i["severity"] == "warning"],
    }


# ── POST /api/validate-excel ──

@router.post("/validate-excel")
def validate_excel(req: ValidateExcelRequest):
    issues = []
    path = Path(req.input_path)

    if not path.exists():
        return {"valid": False, "issues": [_issue("error", "excel", "FILE_NOT_FOUND", f"Excel 文件不存在: {req.input_path}")], "available_sheets": [], "sheet_columns": {}, "warnings": []}

    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"valid": False, "issues": [_issue("error", "excel", "UNREADABLE", f"无法读取 Excel 文件: {e}")], "available_sheets": [], "sheet_columns": {}, "warnings": []}

    available_sheets = list(wb.sheetnames)
    sheet_columns = {}
    config = req.config

    for idx, sheet_cfg in enumerate(config.get("sheets", [])):
        sheet_name = sheet_cfg.get("sheet_name", "")
        if not sheet_name:
            issues.append(_issue("error", "config", "MISSING_SHEET_NAME", f"第 {idx + 1} 个 Sheet 配置缺少 sheet_name", field=f"sheets[{idx}].sheet_name"))
            continue

        if sheet_name not in available_sheets:
            issues.append(_issue("error", "excel", "MISSING_SHEET", f"Excel 中找不到 Sheet '{sheet_name}'", sheet=sheet_name, suggestion=f"可用的 Sheet: {', '.join(available_sheets)}"))
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=5, values_only=True))
        header_row_idx = sheet_cfg.get("header_row", 1) - 1
        if header_row_idx < len(rows):
            headers = [str(c).strip() for c in rows[header_row_idx] if c is not None]
            sheet_columns[sheet_name] = headers

            for field_key, col_label in sheet_cfg.get("columns", {}).items():
                if col_label not in headers:
                    issues.append(_issue("warning", "excel", "MISSING_COLUMN", f"Sheet '{sheet_name}' 中找不到列 '{col_label}'（字段 '{field_key}'）", sheet=sheet_name, column=col_label))

    # Summary sheet
    summary = config.get("summary")
    if summary and summary.get("sheet_name"):
        sn = summary["sheet_name"]
        if sn not in available_sheets:
            issues.append(_issue("warning", "excel", "MISSING_SUMMARY_SHEET", f"汇总页 Sheet '{sn}' 不存在", sheet=sn, suggestion=f"可用的 Sheet: {', '.join(available_sheets)}"))

    wb.close()
    has_errors = any(i["severity"] == "error" for i in issues)
    return {
        "valid": not has_errors,
        "issues": issues,
        "available_sheets": available_sheets,
        "sheet_columns": sheet_columns,
        "warnings": [i["message"] for i in issues if i["severity"] == "warning"],
    }


# ── POST /api/validate-config ──

@router.post("/validate-config")
def validate_config(req: ValidateConfigRequest):
    issues = []
    cfg = req.config

    if not cfg.get("title", "").strip():
        issues.append(_issue("error", "config", "MISSING_TITLE", "配置缺少报告标题", field="title"))

    sheets = cfg.get("sheets", [])
    if not sheets:
        issues.append(_issue("warning", "config", "NO_SHEETS", "配置没有定义任何数据工作表", field="sheets"))

    seen_ids = set()
    for i, sheet in enumerate(sheets):
        name = sheet.get("name", f"第{i+1}个")
        for field in ["name", "sheet_name", "id"]:
            if not sheet.get(field, "").strip():
                issues.append(_issue("error", "config", "MISSING_FIELD", f"工作表 '{name}' 缺少必填字段: {field}", field=f"sheets[{i}].{field}"))

        sid = sheet.get("id", "")
        if sid and sid in seen_ids:
            issues.append(_issue("error", "config", "DUPLICATE_ID", f"重复的工作表 id: {sid}", field=f"sheets[{i}].id"))
        if sid:
            seen_ids.add(sid)

        columns = sheet.get("columns", {})
        if not columns:
            issues.append(_issue("warning", "config", "NO_COLUMNS", f"工作表 '{name}' 没有列映射", field=f"sheets[{i}].columns"))

    summary = cfg.get("summary")
    if summary and summary.get("mode") == "table":
        if not summary.get("key_column"):
            issues.append(_issue("warning", "config", "MISSING_KEY_COLUMN", "汇总模式 'table' 需要设置 key_column", field="summary.key_column"))
        if not summary.get("value_column"):
            issues.append(_issue("warning", "config", "MISSING_VALUE_COLUMN", "汇总模式 'table' 需要设置 value_column", field="summary.value_column"))

    has_errors = any(i["severity"] == "error" for i in issues)
    return {
        "valid": not has_errors,
        "issues": issues,
        "warnings": [i["message"] for i in issues if i["severity"] == "warning"],
    }


# ── POST /api/cross-validate ──

@router.post("/cross-validate")
def cross_validate(req: CrossValidateRequest):
    issues = []

    # Template structure
    try:
        from report_engine.template_parser import parse_template
        structure, _ = parse_template(req.template_path)
    except Exception as e:
        issues.append(_issue("error", "template", "TEMPLATE_UNREADABLE", f"模板解析失败: {e}"))
        return {"valid": False, "issues": issues, "warnings": []}

    template_sections = {s.get("id"): s for s in structure.get("sections", [])} if structure else {}

    # Excel sheets
    try:
        from openpyxl import load_workbook
        wb = load_workbook(req.excel_path, read_only=True)
        excel_sheets = set(wb.sheetnames)
        wb.close()
    except Exception as e:
        issues.append(_issue("error", "excel", "EXCEL_UNREADABLE", f"Excel 读取失败: {e}"))
        return {"valid": False, "issues": issues, "warnings": []}

    # Config sheets vs Excel
    for i, sheet_cfg in enumerate(req.config.get("sheets", [])):
        sn = sheet_cfg.get("sheet_name", "")
        if sn and sn not in excel_sheets:
            issues.append(_issue("error", "cross_validation", "CONFIG_SHEET_NOT_IN_EXCEL", f"配置引用的 Sheet '{sn}' 在 Excel 中不存在", sheet=sn, field=f"sheets[{i}].sheet_name", suggestion=f"可用的 Sheet: {', '.join(sorted(excel_sheets))}"))

    # Summary sheet
    summary = req.config.get("summary")
    if summary and summary.get("sheet_name") and summary["sheet_name"] not in excel_sheets:
        issues.append(_issue("warning", "cross_validation", "SUMMARY_SHEET_NOT_IN_EXCEL", f"配置汇总页引用的 Sheet '{summary['sheet_name']}' 在 Excel 中不存在", sheet=summary["sheet_name"]))

    # Config sections vs template
    config_section_ids = {s.get("id") for s in req.config.get("sheets", []) if s.get("id")}
    for sec_id, sec in template_sections.items():
        if sec_id not in config_section_ids:
            issues.append(_issue("info", "cross_validation", "TEMPLATE_SECTION_NO_CONFIG", f"模板章节 '{sec.get('title', sec_id)}' 在配置中没有对应的 Sheet 映射", placeholder=sec.get("placeholder")))

    has_errors = any(i["severity"] == "error" for i in issues)
    return {
        "valid": not has_errors,
        "issues": issues,
        "warnings": [i["message"] for i in issues if i["severity"] == "warning"],
    }
