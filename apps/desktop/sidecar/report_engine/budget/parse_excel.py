#!/usr/bin/env python3
"""从预算 Excel 提取数据和图片，生成 report-engine 简化内容描述。

用法:
    python parse_excel_budget.py \
        --input budget.xlsx \
        --output-dir output/budget/ \
        --config config.json

配置文件格式:
    {
      "title": "XX项目预算报告",
      "summary": {
        "sheet_name": "汇总页",
        "mode": "table",
        "header_row": 1,
        "key_column": "科目",
        "value_column": "金额（元）",
        "prefix": "SUMMARY_"
      },
      "sheets": [
        {
          "name": "设备费明细",
          "sheet_name": "设备费",
          "columns": { ... },
          "image_columns": ["报价截图"]
        }
      ]
    }

summary 模式说明:
  - mode="table": 按表格读取，指定 key_column（行标识列）和 value_column（值列）
    结果: SUMMARY_设备费=50000, SUMMARY_合计=93000
  - mode="cell_map": 直接指定单元格 → context key
    {"mappings": {"TOTAL_AMOUNT": "C5", "PROJECT_NAME": "B2"}}
"""

import argparse
import json
import logging
import re
import sys
import warnings
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# openpyxl 对部分 XML 会发出警告，先屏蔽
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"错误: 需要安装 openpyxl。pip install openpyxl\n{e}", file=sys.stderr)
    sys.exit(1)

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None

logger = logging.getLogger("parse_excel_budget")


# ── 工具函数 ─────────────────────────────────────────


def _snake_case(text: str) -> str:
    """将中文/英文名称转为 snake_case id。"""
    # 去掉中文数字前缀
    text = re.sub(r"^[一二三四五六七八九十]+[、．.\s]*", "", text).strip()
    # 中文转拼音首字母（简化：保留中文字符的 unicode 范围）
    if re.search(r"[\u4e00-\u9fff]", text):
        # 对于纯中文，使用一个简化的映射
        # 实际使用时会通过配置覆盖
        return re.sub(r"[^\w]", "_", text).strip("_").lower()
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def _fmt_amount(value: Any) -> str:
    """格式化金额，保留两位小数。"""
    if value is None or value == "":
        return ""
    try:
        d = Decimal(str(value))
        return str(d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except Exception:
        return str(value)


def _is_empty(value: Any) -> bool:
    """判断值是否为空（None、空字符串、仅空白字符）。"""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _safe_value(cell_value: Any) -> str:
    """安全地获取单元格字符串值。"""
    if cell_value is None:
        return ""
    if isinstance(cell_value, (int, float)):
        return str(cell_value)
    return str(cell_value).strip()


# ── 公式计算回退 ─────────────────────────────────────────


def _cell_to_num(value: Any) -> Optional[float]:
    """将单元格值转为数字，非数字返回 None。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def _replace_sum_in_formula(expr: str, sheet) -> str:
    """将 Excel SUM(...) 替换为 Python sum([...])。"""
    pattern = r"SUM\(([^)]+)\)"

    def replacer(match):
        args = match.group(1)
        values = []
        for part in args.split(","):
            part = part.strip()
            if ":" in part:
                # 范围引用如 A1:A5
                start, end = part.split(":")
                try:
                    for row in sheet[start:end]:
                        for cell in row:
                            v = _cell_to_num(cell.value)
                            if v is not None:
                                values.append(v)
                except Exception:
                    pass
            else:
                # 单个单元格
                try:
                    v = _cell_to_num(sheet[part].value)
                    if v is not None:
                        values.append(v)
                except Exception:
                    pass
        return f"sum([{','.join(str(v) for v in values)}])"

    return re.sub(pattern, replacer, expr, flags=re.IGNORECASE)


def _replace_cell_refs_in_formula(expr: str, sheet) -> str:
    """将公式中的单元格引用（如 A1, B2）替换为实际数值。"""
    # 匹配大写字母+数字的单元格引用（排除已替换的 sum 等）
    pattern = r"\b([A-Z]{1,3})(\d{1,7})\b"

    def replacer(match):
        cell_ref = match.group(0)
        try:
            val = sheet[cell_ref].value
            num = _cell_to_num(val)
            if num is not None:
                return str(num)
            if val is not None:
                # 文本值用引号包裹
                return repr(str(val))
            return "0"
        except Exception:
            return "0"

    return re.sub(pattern, replacer, expr)


def _evaluate_formula(formula: str, sheet) -> Optional[str]:
    """尝试用 Python 计算 Excel 公式，返回字符串结果或 None。

    支持的公式:
    - =SUM(A1:A5) / =SUM(A1,A2,A3)
    - =A1+B2, =A1-B2, =A1*B2, =A1/B2
    - 混合运算如 =A1+B2*C3
    """
    if not formula or not formula.startswith("="):
        return None

    expr = formula[1:]  # 去掉 '='

    try:
        # 先替换 SUM 函数
        expr = _replace_sum_in_formula(expr, sheet)
        # 再替换剩余单元格引用
        expr = _replace_cell_refs_in_formula(expr, sheet)

        # 安全 eval（仅允许 sum 和基本运算符）
        result = eval(expr, {"__builtins__": {}, "sum": sum}, {})
        if result is None:
            return None
        # 格式化数字（整数去掉小数点）
        if isinstance(result, float):
            if result == int(result):
                return str(int(result))
            return f"{result:.2f}".rstrip("0").rstrip(".")
        return str(result)
    except Exception as e:
        logger.debug("公式计算失败 '%s' -> '%s': %s", formula, expr, e)
        return None


def _get_cell_value(data_cell, formula_cell) -> str:
    """获取单元格值，支持公式自动计算回退。

    优先使用 data_only=True 的缓存值；如果为空且 formula_cell 是公式，
    则尝试用 Python 计算。
    """
    # 1. 尝试 data_only 值
    value = data_cell.value
    if value is not None and str(value).strip() != "":
        if isinstance(value, (int, float)):
            return str(value)
        return str(value).strip()

    # 2. 检查是否是公式，尝试计算
    formula = formula_cell.value
    if isinstance(formula, str) and formula.startswith("="):
        sheet = formula_cell.parent
        calculated = _evaluate_formula(formula, sheet)
        if calculated is not None:
            return calculated

    # 3. 返回空
    return ""


# ── 汇总页读取 ─────────────────────────────────────────


def _read_summary_table(sheet, formula_sheet, summary_config: dict) -> Dict[str, str]:
    """以 table 模式读取汇总页，返回 {context_key: value} 字典。"""
    result = {}
    header_row = summary_config.get("header_row", 1)
    key_col_name = summary_config.get("key_column")
    value_col_name = summary_config.get("value_column")
    prefix = summary_config.get("prefix", "SUMMARY_")

    if not key_col_name or not value_col_name:
        logger.warning("summary table 模式需要 key_column 和 value_column")
        return result

    # 读取表头
    header_values = []
    for cell in sheet[header_row]:
        header_values.append(_safe_value(cell.value))

    key_col_idx = None
    value_col_idx = None
    for idx, hv in enumerate(header_values):
        if hv == key_col_name:
            key_col_idx = idx
        if hv == value_col_name:
            value_col_idx = idx

    if key_col_idx is None:
        logger.warning("汇总页未找到 key_column '%s'，可用: %s", key_col_name, header_values)
        return result
    if value_col_idx is None:
        logger.warning("汇总页未找到 value_column '%s'，可用: %s", value_col_name, header_values)
        return result

    # 读取数据行（同时遍历 data_sheet 和 formula_sheet）
    data_iter = sheet.iter_rows(min_row=header_row + 1, values_only=False)
    formula_iter = formula_sheet.iter_rows(min_row=header_row + 1, values_only=False) if formula_sheet else None
    row_pairs = zip(data_iter, formula_iter) if formula_iter else ((r, None) for r in data_iter)

    for row_idx, (data_row, formula_row) in enumerate(row_pairs, start=header_row + 1):
        data_key_cell = data_row[key_col_idx] if key_col_idx < len(data_row) else None
        formula_key_cell = formula_row[key_col_idx] if formula_row and key_col_idx < len(formula_row) else None
        key = _get_cell_value(data_key_cell, formula_key_cell)

        data_value_cell = data_row[value_col_idx] if value_col_idx < len(data_row) else None
        formula_value_cell = formula_row[value_col_idx] if formula_row and value_col_idx < len(formula_row) else None
        value = _get_cell_value(data_value_cell, formula_value_cell)

        if not key:
            continue

        # 生成 context key: prefix + snake_case(key)
        context_key = f"{prefix}{_snake_case(key)}"
        # 对金额类值格式化
        if _looks_like_amount(value):
            value = _fmt_amount(value)

        result[context_key] = value
        logger.debug("汇总页映射: %s -> %s", context_key, value)

    return result


def _read_summary_cell_map(sheet, formula_sheet, summary_config: dict) -> Dict[str, str]:
    """以 cell_map 模式读取汇总页，返回 {context_key: value} 字典。"""
    result = {}
    mappings = summary_config.get("mappings", {})

    for context_key, cell_ref in mappings.items():
        try:
            data_cell = sheet[cell_ref]
            formula_cell = formula_sheet[cell_ref] if formula_sheet else None
            value = _get_cell_value(data_cell, formula_cell)
            if _looks_like_amount(value):
                value = _fmt_amount(value)
            result[context_key] = value
            logger.debug("汇总页单元格映射: %s (%s) -> %s", context_key, cell_ref, value)
        except Exception as e:
            logger.warning("读取汇总页单元格 %s 失败: %s", cell_ref, e)

    return result


def _looks_like_amount(value: str) -> bool:
    """判断字符串是否像金额（纯数字或带小数点的数字）。"""
    if not value:
        return False
    return bool(re.match(r"^[\d,]+(\.\d+)?$", value.strip().replace(",", "")))


def _read_summary_sheet(wb_data, wb_formula, summary_config: dict) -> Dict[str, str]:
    """读取汇总页，返回 context 变量字典。"""
    sheet_name = summary_config.get("sheet_name")
    if not sheet_name:
        return {}

    if sheet_name not in wb_data.sheetnames:
        logger.warning("汇总页 sheet '%s' 不存在，可用: %s", sheet_name, wb_data.sheetnames)
        return {}

    sheet = wb_data[sheet_name]
    formula_sheet = wb_formula[sheet_name] if wb_formula and sheet_name in wb_formula.sheetnames else None
    mode = summary_config.get("mode", "table")
    logger.info("读取汇总页: %s (mode=%s)", sheet_name, mode)

    if mode == "cell_map":
        return _read_summary_cell_map(sheet, formula_sheet, summary_config)
    else:
        return _read_summary_table(sheet, formula_sheet, summary_config)


# ── 图片提取 ─────────────────────────────────────────


def _extract_images_from_sheet(sheet) -> List[Dict[str, Any]]:
    """从 worksheet 提取所有图片，返回包含 row/col/数据 的字典列表。

    支持两种方式：
    1. openpyxl 3.1+ 的 sheet._images（嵌入图片）
    2. 通过 sheet._drawing 解析（浮动图片）
    """
    images = []

    # 方式1: sheet._images (openpyxl 3.1+)
    if hasattr(sheet, "_images") and sheet._images:
        for img in sheet._images:
            try:
                anchor = img.anchor
                # anchor 可能是 OneCellAnchor 或 TwoCellAnchor
                from_col = getattr(getattr(anchor, "_from", None), "col", None)
                from_row = getattr(getattr(anchor, "_from", None), "row", None)

                if from_col is None and hasattr(anchor, "from_"):
                    from_col = getattr(anchor.from_, "col", None)
                    from_row = getattr(anchor.from_, "row", None)

                if from_row is None or from_col is None:
                    continue

                # 获取图片数据
                img_data = _get_image_data(img)
                if img_data:
                    images.append({
                        "row": from_row,
                        "col": from_col,
                        "data": img_data,
                        "source": "_images",
                    })
            except Exception as e:
                logger.warning("解析嵌入图片失败: %s", e)

    # 方式2: 通过 drawing/rels 解析
    try:
        wb = sheet.parent
        if hasattr(sheet, "_drawing") and sheet._drawing:
            drawing = sheet._drawing
            # 获取 rels 中的图片关系
            if hasattr(drawing, "_rels") and drawing._rels:
                for rel in drawing._rels.values():
                    if "image" in rel.reltype:
                        # 从 workbook 的 _package 中获取图片数据
                        img_data = _get_image_from_rels(wb, rel.target_ref)
                        if img_data:
                            # 尝试找到图片对应的 anchor
                            for anchor_info in _find_anchors_for_image(drawing, rel.rId):
                                images.append({
                                    "row": anchor_info["row"],
                                    "col": anchor_info["col"],
                                    "data": img_data,
                                    "source": "drawing",
                                })
    except Exception as e:
        logger.warning("解析浮动图片失败: %s", e)

    return images


def _get_image_data(img) -> Optional[bytes]:
    """从 openpyxl Image 对象获取二进制数据。"""
    try:
        # 方式1: 直接访问 _data()
        if hasattr(img, "_data"):
            return img._data()
    except Exception:
        pass

    try:
        # 方式2: 访问 ref
        if hasattr(img, "ref"):
            return bytes(img.ref)
    except Exception:
        pass

    try:
        # 方式3: 访问 embed
        if hasattr(img, "embed") and img.embed:
            return bytes(img.embed)
    except Exception:
        pass

    return None


def _get_image_from_rels(wb, target_ref: str) -> Optional[bytes]:
    """从 workbook 的 package 中通过 target_ref 获取图片数据。"""
    try:
        if hasattr(wb, "_package") and wb._package:
            pkg = wb._package
            if hasattr(pkg, "images"):
                for img in pkg.images:
                    if hasattr(img, "path") and target_ref in str(img.path):
                        return img.data if hasattr(img, "data") else None
    except Exception:
        pass
    return None


def _find_anchors_for_image(drawing, rel_id: str) -> List[Dict[str, int]]:
    """在 drawing 中查找指定 rel_id 的图片的锚点坐标。"""
    anchors = []
    try:
        if hasattr(drawing, "spTree") and drawing.spTree:
            for elem in drawing.spTree:
                # 查找包含该图片引用的元素
                blip = elem.find(".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip")
                if blip is not None:
                    embed = blip.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
                    if embed == rel_id:
                        # 查找锚点
                        from_elem = elem.find(".//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from")
                        if from_elem is not None:
                            col_elem = from_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}col")
                            row_elem = from_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row")
                            if col_elem is not None and row_elem is not None:
                                anchors.append({
                                    "col": int(col_elem.text) if col_elem.text else 0,
                                    "row": int(row_elem.text) if row_elem.text else 0,
                                })
    except Exception:
        pass
    return anchors


def _save_image(img_data: bytes, output_path: Path) -> bool:
    """保存图片数据到文件，自动处理格式转换。"""
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # 尝试用 Pillow 检测和转换格式
        if PILImage is not None:
            from io import BytesIO

            try:
                pil_img = PILImage.open(BytesIO(img_data))
                # 统一转为 PNG
                if pil_img.mode in ("RGBA", "P"):
                    pil_img.save(output_path, "PNG")
                else:
                    pil_img.save(output_path, "PNG")
                return True
            except Exception:
                pass

        # 直接保存为 png
        output_path.write_bytes(img_data)
        return True
    except Exception as e:
        logger.warning("保存图片失败: %s", e)
        return False


# ── 数据读取 ─────────────────────────────────────────


def _read_sheet_data(sheet, formula_sheet, config: dict, output_dir: Path) -> Tuple[List[Dict[str, Any]], List[str]]:
    """读取一个 sheet 的数据，返回 (数据行列表, 警告列表)。

    formula_sheet 用于公式计算回退：当 data_only=True 读取的缓存值为空时，
    尝试用 Python 计算公式。
    """
    columns_config = config.get("columns", {})
    image_columns = config.get("image_columns", [])
    header_row = config.get("header_row", 1)  # openpyxl 是 1-based

    warnings_list = []

    # 读取表头
    header_values = []
    for cell in sheet[header_row]:
        header_values.append(_safe_value(cell.value))

    # 建立列映射: config_key -> column_index (0-based)
    col_map: Dict[str, int] = {}
    for key, header_name in columns_config.items():
        for idx, hv in enumerate(header_values):
            if hv == header_name:
                col_map[key] = idx
                break
        if key not in col_map:
            warnings_list.append(f"未找到列 '{header_name}'（对应字段 '{key}'）")

    # 建立图片列映射
    image_col_indices: set = set()
    for img_col_name in image_columns:
        for idx, hv in enumerate(header_values):
            if hv == img_col_name:
                image_col_indices.add(idx)
                break

    # 提取所有图片
    all_images = _extract_images_from_sheet(sheet)
    logger.info("Sheet '%s' 提取到 %d 张图片", sheet.title, len(all_images))

    # 按 (row, col) 索引图片
    images_by_cell: Dict[Tuple[int, int], List[bytes]] = {}
    for img_info in all_images:
        key = (img_info["row"], img_info["col"])
        if key not in images_by_cell:
            images_by_cell[key] = []
        images_by_cell[key].append(img_info["data"])

    # 按行收集图片
    images_by_row: Dict[int, List[bytes]] = {}
    for (row, col), img_data_list in images_by_cell.items():
        # 将 0-based row 转为数据行号（相对于 header_row）
        data_row = row  # 保持原始坐标
        if data_row not in images_by_row:
            images_by_row[data_row] = []
        images_by_row[data_row].extend(img_data_list)

    # 读取数据行（同时遍历 data_sheet 和 formula_sheet）
    data_iter = sheet.iter_rows(min_row=header_row + 1, values_only=False)
    formula_iter = formula_sheet.iter_rows(min_row=header_row + 1, values_only=False) if formula_sheet else None
    row_pairs = zip(data_iter, formula_iter) if formula_iter else ((r, None) for r in data_iter)

    data_rows = []
    for row_idx, (data_row, formula_row) in enumerate(row_pairs, start=header_row + 1):
        row_data = {"__row_idx__": row_idx}

        # 检查是否为空行（所有映射列都为空）
        all_empty = True
        for key, col_idx in col_map.items():
            data_cell = data_row[col_idx] if col_idx < len(data_row) else None
            formula_cell = formula_row[col_idx] if formula_row and col_idx < len(formula_row) else None
            value = _get_cell_value(data_cell, formula_cell)
            row_data[key] = value
            if not _is_empty(value):
                all_empty = False

        if all_empty:
            continue

        # 收集该行的图片（通过行坐标匹配）
        # openpyxl row 是 1-based，所以 data row 的 row_idx 就是 openpyxl row
        row_images = images_by_row.get(row_idx - 1, [])  # _images 的 row 是 0-based
        if not row_images:
            row_images = images_by_row.get(row_idx, [])  # 再试 1-based

        # 也检查图片列索引匹配的图片
        for img_col_idx in image_col_indices:
            cell_key = (row_idx - 1, img_col_idx)  # 0-based
            if cell_key in images_by_cell:
                for img_data in images_by_cell[cell_key]:
                    if img_data not in row_images:
                        row_images.append(img_data)

        # 保存图片到文件
        image_paths = []
        for img_idx, img_data in enumerate(row_images):
            img_filename = f"{config.get('sheet_name', 'sheet')}_{row_idx}_{img_idx + 1}.png"
            img_path = (Path(output_dir) / "images" / img_filename).resolve()
            if _save_image(img_data, img_path):
                image_paths.append(str(img_path))

        row_data["__image_paths__"] = image_paths

        # 留空检测
        for key in ("name", "reason", "basis"):
            if key in col_map and _is_empty(row_data.get(key)):
                col_name = columns_config.get(key, key)
                warnings_list.append(
                    f"Sheet '{sheet.title}' 第 {row_idx} 行，字段 '{col_name}' 为空"
                )

        data_rows.append(row_data)

    return data_rows, warnings_list


# ── 内容生成 ─────────────────────────────────────────


def _build_table_rows(
    data_rows: List[Dict],
    columns_config: dict,
    table_columns: Optional[List[str]] = None,
) -> Tuple[List[str], List[List[str]]]:
    """构建明细表的 headers 和 rows。

    表格列由 table_columns 参数指定（默认：名称、规格、单价、数量、经费）
    """
    if table_columns is None:
        table_columns = ["name", "spec", "unit_price", "quantity", "amount"]

    headers = []
    for key in table_columns:
        if key in columns_config:
            headers.append(columns_config[key])
        else:
            logger.warning("table_columns 包含未映射的字段 '%s'，可用: %s", key, list(columns_config.keys()))

    rows = []
    for row_data in data_rows:
        row_values = []
        for key in table_columns:
            if key in columns_config:
                value = row_data.get(key, "")
                if key == "amount":
                    value = _fmt_amount(value)
                row_values.append(value)
        rows.append(row_values)

    # 合计行
    total = Decimal("0")
    has_amount = False
    for row_data in data_rows:
        amount_val = row_data.get("amount", "")
        if amount_val:
            try:
                total += Decimal(str(amount_val))
                has_amount = True
            except Exception:
                pass

    if has_amount:
        total_row = ["合计"] + [""] * (len(headers) - 1)
        total_row[-1] = _fmt_amount(total)
        rows.append(total_row)

    return headers, rows


def _strip_chinese_number_prefix(text: str) -> str:
    """去掉中文数字前缀（如 '一、' '十二、' '1.' '1、' 等），用于模板已有自动编号的场景。"""
    return re.sub(r"^[一二三四五六七八九十百]+[、．.\s]*", "", text).strip()


def _build_section(
    data_rows: List[Dict],
    config: dict,
    columns_config: dict,
    table_columns: Optional[List[str]] = None,
    detail_fields: Optional[List[Dict[str, str]]] = None,
    image_columns: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """构建一个 sheet 对应的 section（简化内容描述格式）。"""
    sheet_name = config.get("name", config.get("sheet_name", "明细"))
    section_id = config.get("id") or _snake_case(sheet_name)

    if detail_fields is None:
        detail_fields = [
            {"field": "reason", "label": "购置理由"},
            {"field": "basis", "label": "测算依据"},
        ]
    if image_columns is None:
        image_columns = ["报价截图"]

    heading_level = config.get("heading_level", 2)
    item_heading_level = config.get("item_heading_level", 3)

    # 去掉中文编号前缀（模板已自带自动编号）
    display_name = _strip_chinese_number_prefix(sheet_name)

    blocks = []

    # 1. 科目标题（level=0 时用正文段落）
    if heading_level == 0:
        blocks.append({"type": "paragraph", "text": display_name})
    else:
        blocks.append({"type": "heading", "text": display_name, "level": heading_level})

    # 2. 明细表
    headers, rows = _build_table_rows(data_rows, columns_config, table_columns)
    if headers:
        blocks.append({
            "type": "table",
            "title": f"{display_name}一览",
            "headers": headers,
            "rows": rows,
        })

    # 3. 逐条详情
    for idx, row_data in enumerate(data_rows, start=1):
        name = row_data.get("name", f"项目{idx}")
        image_paths = row_data.get("__image_paths__", [])

        # 明细条目标题（level=0 时用正文段落，带编号；heading 由模板样式自动编号）
        if item_heading_level == 0:
            blocks.append({"type": "paragraph", "text": f"{idx}. {name}"})
        else:
            blocks.append({"type": "heading", "text": name, "level": item_heading_level})

        # 动态字段段落
        for field_def in detail_fields:
            field_key = field_def["field"]
            label = field_def["label"]
            value = row_data.get(field_key, "")
            if _is_empty(value):
                blocks.append({"type": "paragraph", "text": f"{label}：[未填写]"})
            else:
                blocks.append({"type": "paragraph", "text": f"{label}：{value}"})

        # 报价截图（按顺序插入多张）
        if image_columns:
            if image_paths:
                for img_idx, img_path in enumerate(image_paths, start=1):
                    caption = f"报价截图 {img_idx}" if len(image_paths) > 1 else "报价截图"
                    blocks.append({
                        "type": "image",
                        "path": img_path,
                        "caption": caption,
                        "width_cm": 14,
                    })
            else:
                blocks.append({
                    "type": "paragraph",
                    "text": "报价截图：[未上传]",
                })

    return {
        "name": sheet_name,
        "id": section_id,
        "blocks": blocks,
    }


def parse_excel_budget(input_path: str, output_dir: str, config: dict) -> Tuple[Dict[str, Any], List[str]]:
    """主函数：解析 Excel 并生成简化内容描述。"""
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    all_warnings = []

    # 加载工作簿（data_only=True 读取缓存值，data_only=False 用于公式计算回退）
    logger.info("加载 Excel: %s", input_path)
    wb_data = load_workbook(input_path, data_only=True)
    wb_formula = load_workbook(input_path, data_only=False)

    sections = []
    for sheet_config in config.get("sheets", []):
        sheet_name = sheet_config.get("sheet_name")
        if not sheet_name:
            continue

        if sheet_name not in wb_data.sheetnames:
            msg = f"Sheet '{sheet_name}' 不存在，可用: {wb_data.sheetnames}"
            logger.warning(msg)
            all_warnings.append(msg)
            continue

        sheet = wb_data[sheet_name]
        formula_sheet = wb_formula[sheet_name]
        logger.info("处理 Sheet: %s", sheet_name)

        data_rows, warnings = _read_sheet_data(sheet, formula_sheet, sheet_config, output_dir)
        all_warnings.extend(warnings)

        if not data_rows:
            logger.warning("Sheet '%s' 未读取到数据", sheet_name)
            continue

        section = _build_section(
            data_rows,
            sheet_config,
            sheet_config.get("columns", {}),
            table_columns=sheet_config.get("table_columns"),
            detail_fields=sheet_config.get("detail_fields"),
            image_columns=sheet_config.get("image_columns"),
        )
        sections.append(section)

        logger.info("Sheet '%s' 读取 %d 行数据，%d 张图片", sheet_name, len(data_rows),
                    sum(len(r.get("__image_paths__", [])) for r in data_rows))

    # 读取汇总页数据作为 extra_context
    extra_context = {}
    if "summary" in config:
        summary_data = _read_summary_sheet(wb_data, wb_formula, config["summary"])
        extra_context.update(summary_data)

    content = {
        "title": config.get("title", "预算报告"),
        "sections": sections,
        "attachments": [],
        "extra_context": extra_context,
    }

    return content, all_warnings


def main():
    parser = argparse.ArgumentParser(description="从预算 Excel 提取数据和图片，生成 report-engine 内容描述")
    parser.add_argument("--input", required=True, help="Excel 文件路径")
    parser.add_argument("--output-dir", required=True, help="输出目录")
    parser.add_argument("--config", required=True, help="配置文件路径 (JSON)")
    parser.add_argument("--verbose", "-v", action="store_true", help="详细日志")
    parser.add_argument("--dry-run", action="store_true", help="仅预览数据结构，不保存图片")
    args = parser.parse_args()

    # 配置日志
    level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    # 读取配置
    config = json.loads(Path(args.config).read_text(encoding="utf-8"))

    # 解析
    content, warnings_list = parse_excel_budget(args.input, args.output_dir, config)

    # 输出警告
    if warnings_list:
        logger.warning("=" * 50)
        logger.warning("解析警告 (%d 条):", len(warnings_list))
        for w in warnings_list:
            logger.warning("  - %s", w)
        logger.warning("=" * 50)

    # 预览或保存
    if args.dry_run:
        print("\n=== 数据结构预览 ===")
        # 简化输出，去掉图片路径的完整内容
        preview = json.loads(json.dumps(content, ensure_ascii=False))
        for sec in preview.get("sections", []):
            for blk in sec.get("blocks", []):
                if blk.get("type") == "image":
                    blk["path"] = "..."
        print(json.dumps(preview, ensure_ascii=False, indent=2))
    else:
        content_path = Path(args.output_dir) / "content.json"
        content_path.write_text(
            json.dumps(content, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        logger.info("内容描述已保存: %s", content_path)
        logger.info("  sections: %d", len(content.get("sections", [])))

    # 如果有警告，以非零退出码退出（但内容已保存）
    if warnings_list:
        logger.warning("存在 %d 条警告，请检查日志", len(warnings_list))


if __name__ == "__main__":
    main()
