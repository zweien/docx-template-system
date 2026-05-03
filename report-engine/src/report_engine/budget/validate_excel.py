"""Excel 数据校验：根据 BudgetConfig 校验 Excel 文件内容完整性。"""

import logging
from typing import Any, Optional

import openpyxl

from report_engine.budget.models import (
    ExcelValidationResponse,
    SheetResult,
    SummaryResult,
)

logger = logging.getLogger(__name__)


# ── 复用 parse_excel_budget 的辅助函数 ──

def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _safe_value(cell_value: Any) -> str:
    if cell_value is None:
        return ""
    if isinstance(cell_value, (int, float)):
        return str(cell_value)
    return str(cell_value).strip()


def _cell_to_num(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


# ── 数值型字段关键词 ──

_NUMERIC_KEYWORDS = ("price", "amount", "quantity", "金额", "费用", "数量", "单价", "经费")


def _is_numeric_field(field: str) -> bool:
    fl = field.lower()
    return any(kw in fl for kw in _NUMERIC_KEYWORDS)


# ── 图片统计（轻量版，不提取二进制数据）──

def _count_images_in_sheet(sheet) -> dict:
    """统计 sheet 中的图片数量和行分布（不读取图片数据）。"""
    total = 0
    rows: set[int] = set()

    if hasattr(sheet, "_images") and sheet._images:
        for img in sheet._images:
            try:
                anchor = img.anchor
                from_row = getattr(getattr(anchor, "_from", None), "row", None)
                if from_row is None and hasattr(anchor, "from_"):
                    from_row = getattr(anchor.from_, "row", None)
                if from_row is not None:
                    rows.add(from_row)
                total += 1
            except Exception:
                total += 1

    try:
        if hasattr(sheet, "_drawing") and sheet._drawing:
            drawing = sheet._drawing
            if hasattr(drawing, "_rels") and drawing._rels:
                for rel in drawing._rels.values():
                    if "image" in rel.reltype:
                        try:
                            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
                            anchors = _find_anchors_for_image(drawing, rel.rId)
                            for a in anchors:
                                rows.add(a["row"])
                            total += max(len(anchors), 1)
                        except Exception:
                            total += 1
    except Exception:
        pass

    return {"total_images": total, "rows_with_images": len(rows)}


def _find_anchors_for_image(drawing, rel_id: str) -> list[dict]:
    anchors = []
    try:
        for sp in (drawing._spTree or []):
            if hasattr(sp, "nvSpPr") and hasattr(sp.nvSpPr, "cNvPr"):
                desc = getattr(sp.nvSpPr.cNvPr, "descr", "") or ""
                if not desc:
                    continue
            bv = None
            if hasattr(sp, "spPr") and hasattr(sp.spPr, "xfrm"):
                xfrm = sp.spPr.xfrm
                if hasattr(xfrm, "off") and xfrm.off:
                    row = getattr(xfrm.off, "row", 0)
                    col = getattr(xfrm.off, "col", 0)
                    anchors.append({"row": row, "col": col})
    except Exception:
        pass
    return anchors


# ── 校验逻辑 ──

def _validate_detail_sheet(
    sheet,
    sheet_cfg: dict,
) -> SheetResult:
    """校验单个 detail sheet。"""
    result = SheetResult(sheet_name=sheet_cfg.get("sheet_name", ""), found=True)
    columns = sheet_cfg.get("columns", {})
    image_columns = sheet_cfg.get("image_columns", [])

    # 读取表头行（默认第 1 行）
    header_row_idx = sheet_cfg.get("header_row", 1)
    header_cells = list(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=False))
    if not header_cells:
        result.warnings.append("无法读取表头行")
        return result

    header_values = [_safe_value(c.value) for c in header_cells[0]]

    # 构建列映射：field -> col_index
    col_map: dict[str, int] = {}
    header_set = set(header_values)
    config_headers = set(columns.values())

    for field, header_name in columns.items():
        if header_name in header_values:
            col_map[field] = header_values.index(header_name)
        # header_name 不在 header_values 中则视为缺失

    result.missing_columns = [v for v in config_headers if v not in header_set]
    result.extra_columns = [v for v in header_set if v and v not in config_headers]

    # 逐行扫描数据
    data_start_row = header_row_idx + 1
    total_data_rows = 0
    total_mapped_cells = 0
    non_empty_cells = 0
    value_sets: dict[str, set[str]] = {f: set() for f in columns}

    MAX_ROWS = 10000
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=data_start_row, max_row=data_start_row + MAX_ROWS, values_only=False),
        start=data_start_row,
    ):
        values = [_safe_value(c.value) for c in row]
        # 跳过全空行
        if all(v == "" for v in values):
            continue

        total_data_rows += 1

        for field, col_idx in col_map.items():
            total_mapped_cells += 1
            val = values[col_idx] if col_idx < len(values) else ""

            if _is_empty(val):
                result.empty_cells.append({
                    "row": row_idx,
                    "column": columns[field],
                    "field": field,
                })
            else:
                non_empty_cells += 1
                # 截断过长的值
                val_str = val[:100] if len(val) > 100 else val
                value_sets[field].add(val_str)

                # 数值类型检查
                if _is_numeric_field(field) and _cell_to_num(val) is None:
                    result.numeric_violations.append({
                        "row": row_idx,
                        "column": columns[field],
                        "value": val[:50],
                    })

    result.total_rows = total_data_rows
    result.fill_rate = round(non_empty_cells / total_mapped_cells, 4) if total_mapped_cells > 0 else 0.0
    result.unique_values = {f: sorted(vs)[:20] for f, vs in value_sets.items()}

    # 图片统计
    if image_columns:
        result.image_summary = _count_images_in_sheet(sheet)

    return result


def _validate_summary_sheet(
    wb: openpyxl.Workbook,
    summary_cfg: dict,
) -> SummaryResult:
    """校验汇总 sheet。"""
    sheet_name = summary_cfg.get("sheet_name", "")
    mode = summary_cfg.get("mode", "table")
    result = SummaryResult(sheet_name=sheet_name, found=False, mode=mode)

    if sheet_name not in wb.sheetnames:
        return result

    result.found = True
    sheet = wb[sheet_name]

    if mode == "table":
        header_row = summary_cfg.get("header_row", 1)
        key_column = summary_cfg.get("key_column", "")
        value_column = summary_cfg.get("value_column", "")

        header_cells = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=False))
        if header_cells:
            header_values = [_safe_value(c.value) for c in header_cells[0]]
            result.key_column_found = key_column in header_values
            result.value_column_found = value_column in header_values

            if result.key_column_found and result.value_column_found:
                key_idx = header_values.index(key_column)
                val_idx = header_values.index(value_column)
                mapped = 0
                for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                    k = _safe_value(row[key_idx]) if key_idx < len(row) else ""
                    v = _safe_value(row[val_idx]) if val_idx < len(row) else ""
                    if k and v:
                        mapped += 1
                    elif k and not v:
                        result.missing_keys.append(k)
                result.mapped_count = mapped

    elif mode == "cell_map":
        mappings = summary_cfg.get("mappings", {})
        mapped = 0
        for key, cell_ref in mappings.items():
            try:
                val = _safe_value(sheet[cell_ref].value)
                if val:
                    mapped += 1
                else:
                    result.missing_keys.append(key)
            except Exception:
                result.missing_keys.append(key)
        result.mapped_count = mapped

    return result


def validate_excel_data(input_path: str, config: dict) -> ExcelValidationResponse:
    """主校验函数：校验 Excel 文件是否满足 BudgetConfig 要求。"""
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        return ExcelValidationResponse(
            success=False,
            config_title=config.get("title", ""),
            error={"code": "FILE_ERROR", "message": f"无法读取 Excel 文件: {e}"},
        )

    config_sheets = config.get("sheets", [])
    config_sheet_names = {s.get("sheet_name", "") for s in config_sheets}
    excel_sheets = wb.sheetnames

    missing_sheets = [name for name in config_sheet_names if name and name not in excel_sheets]

    # 校验每个 detail sheet
    sheet_results: list[SheetResult] = []
    total_errors = 0
    total_warnings = 0

    for s_cfg in config_sheets:
        s_name = s_cfg.get("sheet_name", "")
        if not s_name:
            continue
        if s_name in missing_sheets:
            sr = SheetResult(sheet_name=s_name, found=False)
            sr.warnings.append(f"Sheet '{s_name}' 不存在于 Excel 中")
            sheet_results.append(sr)
            total_errors += 1
            continue

        sheet = wb[s_name]
        sr = _validate_detail_sheet(sheet, s_cfg)
        sheet_results.append(sr)

        if sr.missing_columns:
            total_errors += len(sr.missing_columns)
        total_warnings += len(sr.warnings)
        total_warnings += len(sr.numeric_violations)

    # 校验汇总 sheet
    summary_result = None
    summary_cfg = config.get("summary")
    if summary_cfg:
        summary_result = _validate_summary_sheet(wb, summary_cfg)
        if not summary_result.found:
            total_warnings += 1
        total_warnings += len(summary_result.missing_keys)

    wb.close()

    overall_pass = (
        len(missing_sheets) == 0
        and not any(sr.missing_columns for sr in sheet_results)
    )

    return ExcelValidationResponse(
        success=True,
        config_title=config.get("title", ""),
        excel_sheets=excel_sheets,
        missing_sheets=missing_sheets,
        summary=summary_result,
        sheets=sheet_results,
        overall_pass=overall_pass,
        total_errors=total_errors,
        total_warnings=total_warnings,
    )
