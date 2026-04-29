#!/usr/bin/env python3
"""从预算 Excel 提取数据和图片，生成 report-engine 简化内容描述。

用法:
    python parse_excel_budget.py \
        --input budget.xlsx \
        --output-dir output/budget/ \
        --config config.json

配置文件格式:
    {
      "title": "XX项目预算报告",
      "sheets": [
        {
          "name": "设备费明细",
          "sheet_name": "设备费",
          "columns": {
            "name": "名称",
            "spec": "规格",
            "unit_price": "单价",
            "quantity": "数量",
            "amount": "经费",
            "reason": "购置理由",
            "basis": "测算依据"
          },
          "image_columns": ["报价截图"]
        }
      ]
    }
"""

import argparse
import json
import logging
import re
import sys
import warnings
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# openpyxl 对部分 XML 会发出警告，先屏蔽
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"错误: 需要安装 openpyxl。pip install openpyxl\n{e}", file=sys.stderr)
    sys.exit(1)

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None

logger = logging.getLogger("parse_excel_budget")


# ── 工具函数 ─────────────────────────────────────────


def _snake_case(text: str) -> str:
    """将中文/英文名称转为 snake_case id。"""
    # 去掉中文数字前缀
    text = re.sub(r"^[一二三四五六七八九十]+[、．.\s]*", "", text).strip()
    # 中文转拼音首字母（简化：保留中文字符的 unicode 范围）
    if re.search(r"[\u4e00-\u9fff]", text):
        # 对于纯中文，使用一个简化的映射
        # 实际使用时会通过配置覆盖
        return re.sub(r"[^\w]", "_", text).strip("_").lower()
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def _fmt_amount(value: Any) -> str:
    """格式化金额，保留两位小数。"""
    if value is None or value == "":
        return ""
    try:
        d = Decimal(str(value))
        return str(d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except Exception:
        return str(value)


def _is_empty(value: Any) -> bool:
    """判断值是否为空（None、空字符串、仅空白字符）。"""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _safe_value(cell_value: Any) -> str:
    """安全地获取单元格字符串值。"""
    if cell_value is None:
        return ""
    if isinstance(cell_value, (int, float)):
        return str(cell_value)
    return str(cell_value).strip()


# ── 图片提取 ─────────────────────────────────────────


def _extract_images_from_sheet(sheet) -> List[Dict[str, Any]]:
    """从 worksheet 提取所有图片，返回包含 row/col/数据 的字典列表。

    支持两种方式：
    1. openpyxl 3.1+ 的 sheet._images（嵌入图片）
    2. 通过 sheet._drawing 解析（浮动图片）
    """
    images = []

    # 方式1: sheet._images (openpyxl 3.1+)
    if hasattr(sheet, "_images") and sheet._images:
        for img in sheet._images:
            try:
                anchor = img.anchor
                # anchor 可能是 OneCellAnchor 或 TwoCellAnchor
                from_col = getattr(getattr(anchor, "_from", None), "col", None)
                from_row = getattr(getattr(anchor, "_from", None), "row", None)

                if from_col is None and hasattr(anchor, "from_"):
                    from_col = getattr(anchor.from_, "col", None)
                    from_row = getattr(anchor.from_, "row", None)

                if from_row is None or from_col is None:
                    continue

                # 获取图片数据
                img_data = _get_image_data(img)
                if img_data:
                    images.append({
                        "row": from_row,
                        "col": from_col,
                        "data": img_data,
                        "source": "_images",
                    })
            except Exception as e:
                logger.warning("解析嵌入图片失败: %s", e)

    # 方式2: 通过 drawing/rels 解析
    try:
        wb = sheet.parent
        if hasattr(sheet, "_drawing") and sheet._drawing:
            drawing = sheet._drawing
            # 获取 rels 中的图片关系
            if hasattr(drawing, "_rels") and drawing._rels:
                for rel in drawing._rels.values():
                    if "image" in rel.reltype:
                        # 从 workbook 的 _package 中获取图片数据
                        img_data = _get_image_from_rels(wb, rel.target_ref)
                        if img_data:
                            # 尝试找到图片对应的 anchor
                            for anchor_info in _find_anchors_for_image(drawing, rel.rId):
                                images.append({
                                    "row": anchor_info["row"],
                                    "col": anchor_info["col"],
                                    "data": img_data,
                                    "source": "drawing",
                                })
    except Exception as e:
        logger.warning("解析浮动图片失败: %s", e)

    return images


def _get_image_data(img) -> Optional[bytes]:
    """从 openpyxl Image 对象获取二进制数据。"""
    try:
        # 方式1: 直接访问 _data()
        if hasattr(img, "_data"):
            return img._data()
    except Exception:
        pass

    try:
        # 方式2: 访问 ref
        if hasattr(img, "ref"):
            return bytes(img.ref)
    except Exception:
        pass

    try:
        # 方式3: 访问 embed
        if hasattr(img, "embed") and img.embed:
            return bytes(img.embed)
    except Exception:
        pass

    return None


def _get_image_from_rels(wb, target_ref: str) -> Optional[bytes]:
    """从 workbook 的 package 中通过 target_ref 获取图片数据。"""
    try:
        if hasattr(wb, "_package") and wb._package:
            pkg = wb._package
            if hasattr(pkg, "images"):
                for img in pkg.images:
                    if hasattr(img, "path") and target_ref in str(img.path):
                        return img.data if hasattr(img, "data") else None
    except Exception:
        pass
    return None


def _find_anchors_for_image(drawing, rel_id: str) -> List[Dict[str, int]]:
    """在 drawing 中查找指定 rel_id 的图片的锚点坐标。"""
    anchors = []
    try:
        if hasattr(drawing, "spTree") and drawing.spTree:
            for elem in drawing.spTree:
                # 查找包含该图片引用的元素
                blip = elem.find(".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip")
                if blip is not None:
                    embed = blip.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
                    if embed == rel_id:
                        # 查找锚点
                        from_elem = elem.find(".//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from")
                        if from_elem is not None:
                            col_elem = from_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}col")
                            row_elem = from_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row")
                            if col_elem is not None and row_elem is not None:
                                anchors.append({
                                    "col": int(col_elem.text) if col_elem.text else 0,
                                    "row": int(row_elem.text) if row_elem.text else 0,
                                })
    except Exception:
        pass
    return anchors


def _save_image(img_data: bytes, output_path: Path) -> bool:
    """保存图片数据到文件，自动处理格式转换。"""
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # 尝试用 Pillow 检测和转换格式
        if PILImage is not None:
            from io import BytesIO

            try:
                pil_img = PILImage.open(BytesIO(img_data))
                # 统一转为 PNG
                if pil_img.mode in ("RGBA", "P"):
                    pil_img.save(output_path, "PNG")
                else:
                    pil_img.save(output_path, "PNG")
                return True
            except Exception:
                pass

        # 直接保存为 png
        output_path.write_bytes(img_data)
        return True
    except Exception as e:
        logger.warning("保存图片失败: %s", e)
        return False


# ── 数据读取 ─────────────────────────────────────────


def _read_sheet_data(sheet, config: dict, output_dir: Path) -> Tuple[List[Dict[str, Any]], List[str]]:
    """读取一个 sheet 的数据，返回 (数据行列表, 警告列表)。"""
    columns_config = config.get("columns", {})
    image_columns = config.get("image_columns", [])
    header_row = config.get("header_row", 1)  # openpyxl 是 1-based

    warnings_list = []

    # 读取表头
    header_values = []
    for cell in sheet[header_row]:
        header_values.append(_safe_value(cell.value))

    # 建立列映射: config_key -> column_index (0-based)
    col_map: Dict[str, int] = {}
    for key, header_name in columns_config.items():
        for idx, hv in enumerate(header_values):
            if hv == header_name:
                col_map[key] = idx
                break
        if key not in col_map:
            warnings_list.append(f"未找到列 '{header_name}'（对应字段 '{key}'）")

    # 建立图片列映射
    image_col_indices: set = set()
    for img_col_name in image_columns:
        for idx, hv in enumerate(header_values):
            if hv == img_col_name:
                image_col_indices.add(idx)
                break

    # 提取所有图片
    all_images = _extract_images_from_sheet(sheet)
    logger.info("Sheet '%s' 提取到 %d 张图片", sheet.title, len(all_images))

    # 按 (row, col) 索引图片
    images_by_cell: Dict[Tuple[int, int], List[bytes]] = {}
    for img_info in all_images:
        key = (img_info["row"], img_info["col"])
        if key not in images_by_cell:
            images_by_cell[key] = []
        images_by_cell[key].append(img_info["data"])

    # 按行收集图片
    images_by_row: Dict[int, List[bytes]] = {}
    for (row, col), img_data_list in images_by_cell.items():
        # 将 0-based row 转为数据行号（相对于 header_row）
        data_row = row  # 保持原始坐标
        if data_row not in images_by_row:
            images_by_row[data_row] = []
        images_by_row[data_row].extend(img_data_list)

    # 读取数据行
    data_rows = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
        row_data = {"__row_idx__": row_idx}

        # 检查是否为空行（所有映射列都为空）
        all_empty = True
        for key, col_idx in col_map.items():
            cell = row[col_idx] if col_idx < len(row) else None
            value = _safe_value(cell.value if cell else None)
            row_data[key] = value
            if not _is_empty(value):
                all_empty = False

        if all_empty:
            continue

        # 收集该行的图片（通过行坐标匹配）
        # openpyxl row 是 1-based，所以 data row 的 row_idx 就是 openpyxl row
        row_images = images_by_row.get(row_idx - 1, [])  # _images 的 row 是 0-based
        if not row_images:
            row_images = images_by_row.get(row_idx, [])  # 再试 1-based

        # 也检查图片列索引匹配的图片
        for img_col_idx in image_col_indices:
            cell_key = (row_idx - 1, img_col_idx)  # 0-based
            if cell_key in images_by_cell:
                for img_data in images_by_cell[cell_key]:
                    if img_data not in row_images:
                        row_images.append(img_data)

        # 保存图片到文件
        image_paths = []
        for img_idx, img_data in enumerate(row_images):
            img_filename = f"{config.get('sheet_name', 'sheet')}_{row_idx}_{img_idx + 1}.png"
            img_path = output_dir / "images" / img_filename
            if _save_image(img_data, img_path):
                image_paths.append(str(img_path))

        row_data["__image_paths__"] = image_paths

        # 留空检测
        for key in ("name", "reason", "basis"):
            if key in col_map and _is_empty(row_data.get(key)):
                col_name = columns_config.get(key, key)
                warnings_list.append(
                    f"Sheet '{sheet.title}' 第 {row_idx} 行，字段 '{col_name}' 为空"
                )

        data_rows.append(row_data)

    return data_rows, warnings_list


# ── 内容生成 ─────────────────────────────────────────


def _build_table_rows(data_rows: List[Dict], columns_config: dict) -> Tuple[List[str], List[List[str]]]:
    """构建明细表的 headers 和 rows。

    表格列：名称、规格、单价、数量、经费（不含 reason/basis/image）
    """
    table_col_keys = ["name", "spec", "unit_price", "quantity", "amount"]
    headers = []
    for key in table_col_keys:
        if key in columns_config:
            headers.append(columns_config[key])

    rows = []
    for row_data in data_rows:
        row_values = []
        for key in table_col_keys:
            if key in columns_config:
                value = row_data.get(key, "")
                if key == "amount":
                    value = _fmt_amount(value)
                row_values.append(value)
        rows.append(row_values)

    # 合计行
    total = Decimal("0")
    has_amount = False
    for row_data in data_rows:
        amount_val = row_data.get("amount", "")
        if amount_val:
            try:
                total += Decimal(str(amount_val))
                has_amount = True
            except Exception:
                pass

    if has_amount:
        total_row = ["合计"] + [""] * (len(headers) - 2) + [_fmt_amount(total)]
        if len(total_row) < len(headers):
            total_row = ["合计"] + [""] * (len(headers) - 2) + [_fmt_amount(total)]
            # 确保长度匹配
            while len(total_row) < len(headers):
                total_row.insert(1, "")
        rows.append(total_row)

    return headers, rows


def _build_section(data_rows: List[Dict], config: dict, columns_config: dict) -> Dict[str, Any]:
    """构建一个 sheet 对应的 section（简化内容描述格式）。"""
    sheet_name = config.get("name", config.get("sheet_name", "明细"))
    section_id = config.get("id") or _snake_case(sheet_name)

    blocks = []

    # 1. 科目标题
    blocks.append({"type": "heading", "text": sheet_name, "level": 2})

    # 2. 明细表
    headers, rows = _build_table_rows(data_rows, columns_config)
    if headers:
        blocks.append({
            "type": "table",
            "title": f"表1 {sheet_name}一览",
            "headers": headers,
            "rows": rows,
        })

    # 3. 逐条详情
    for idx, row_data in enumerate(data_rows, start=1):
        name = row_data.get("name", f"项目{idx}")
        reason = row_data.get("reason", "")
        basis = row_data.get("basis", "")
        image_paths = row_data.get("__image_paths__", [])

        # 设备名称（三级标题）
        blocks.append({"type": "heading", "text": f"{idx}. {name}", "level": 3})

        # 购置理由
        if _is_empty(reason):
            blocks.append({"type": "paragraph", "text": "购置理由：[未填写]"})
        else:
            blocks.append({"type": "paragraph", "text": f"购置理由：{reason}"})

        # 测算依据
        if _is_empty(basis):
            blocks.append({"type": "paragraph", "text": "测算依据：[未填写]"})
        else:
            blocks.append({"type": "paragraph", "text": f"测算依据：{basis}"})

        # 报价截图（按顺序插入多张）
        if image_paths:
            for img_idx, img_path in enumerate(image_paths, start=1):
                caption = f"报价截图 {img_idx}" if len(image_paths) > 1 else "报价截图"
                blocks.append({
                    "type": "image",
                    "path": img_path,
                    "caption": caption,
                    "width_cm": 14,
                })
        else:
            blocks.append({
                "type": "paragraph",
                "text": "报价截图：[未上传]",
            })

    return {
        "name": sheet_name,
        "id": section_id,
        "blocks": blocks,
    }


def parse_excel_budget(input_path: str, output_dir: str, config: dict) -> Tuple[Dict[str, Any], List[str]]:
    """主函数：解析 Excel 并生成简化内容描述。"""
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    all_warnings = []

    # 加载工作簿
    logger.info("加载 Excel: %s", input_path)
    wb = load_workbook(input_path, data_only=True)

    sections = []
    for sheet_config in config.get("sheets", []):
        sheet_name = sheet_config.get("sheet_name")
        if not sheet_name:
            continue

        if sheet_name not in wb.sheetnames:
            msg = f"Sheet '{sheet_name}' 不存在，可用: {wb.sheetnames}"
            logger.warning(msg)
            all_warnings.append(msg)
            continue

        sheet = wb[sheet_name]
        logger.info("处理 Sheet: %s", sheet_name)

        data_rows, warnings = _read_sheet_data(sheet, sheet_config, output_dir)
        all_warnings.extend(warnings)

        if not data_rows:
            logger.warning("Sheet '%s' 未读取到数据", sheet_name)
            continue

        section = _build_section(data_rows, sheet_config, sheet_config.get("columns", {}))
        sections.append(section)

        logger.info("Sheet '%s' 读取 %d 行数据，%d 张图片", sheet_name, len(data_rows),
                    sum(len(r.get("__image_paths__", [])) for r in data_rows))

    content = {
        "title": config.get("title", "预算报告"),
        "sections": sections,
        "attachments": [],
    }

    return content, all_warnings


def main():
    parser = argparse.ArgumentParser(description="从预算 Excel 提取数据和图片，生成 report-engine 内容描述")
    parser.add_argument("--input", required=True, help="Excel 文件路径")
    parser.add_argument("--output-dir", required=True, help="输出目录")
    parser.add_argument("--config", required=True, help="配置文件路径 (JSON)")
    parser.add_argument("--verbose", "-v", action="store_true", help="详细日志")
    parser.add_argument("--dry-run", action="store_true", help="仅预览数据结构，不保存图片")
    args = parser.parse_args()

    # 配置日志
    level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    # 读取配置
    config = json.loads(Path(args.config).read_text(encoding="utf-8"))

    # 解析
    content, warnings_list = parse_excel_budget(args.input, args.output_dir, config)

    # 输出警告
    if warnings_list:
        logger.warning("=" * 50)
        logger.warning("解析警告 (%d 条):", len(warnings_list))
        for w in warnings_list:
            logger.warning("  - %s", w)
        logger.warning("=" * 50)

    # 预览或保存
    if args.dry_run:
        print("\n=== 数据结构预览 ===")
        # 简化输出，去掉图片路径的完整内容
        preview = json.loads(json.dumps(content, ensure_ascii=False))
        for sec in preview.get("sections", []):
            for blk in sec.get("blocks", []):
                if blk.get("type") == "image":
                    blk["path"] = "..."
        print(json.dumps(preview, ensure_ascii=False, indent=2))
    else:
        content_path = Path(args.output_dir) / "content.json"
        content_path.write_text(
            json.dumps(content, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        logger.info("内容描述已保存: %s", content_path)
        logger.info("  sections: %d", len(content.get("sections", [])))

    # 如果有警告，以非零退出码退出（但内容已保存）
    if warnings_list:
        logger.warning("存在 %d 条警告，请检查日志", len(warnings_list))


if __name__ == "__main__":
    main()
