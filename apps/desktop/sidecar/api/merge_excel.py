"""Excel 合并端点：多个 xlsx 文件按 sheet 合并数据行，保留图片和格式。"""

import shutil
from copy import copy
from io import BytesIO
from pathlib import Path

from fastapi import APIRouter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

router = APIRouter(tags=["merge"])


# ── 请求/响应模型 ──


class MergeInfoRequest(BaseModel):
    file_paths: list[str]


class SheetInfo(BaseModel):
    sheet_name: str
    header_row: list[str]
    data_row_count: int
    has_images: bool
    column_count: int


class FileSheetInfo(BaseModel):
    file_path: str
    file_name: str
    sheets: list[SheetInfo]


class MergeInfoResponse(BaseModel):
    success: bool
    files: list[FileSheetInfo]
    common_sheets: list[str]
    error: dict | None = None


class MergeExcelRequest(BaseModel):
    base_file: str
    source_files: list[str]
    selected_sheets: list[str]
    output_path: str


class SheetMismatchDetail(BaseModel):
    sheet_name: str
    file_name: str
    base_headers: list[str]
    file_headers: list[str]
    missing_in_file: list[str]
    extra_in_file: list[str]


class MergeExcelResponse(BaseModel):
    success: bool
    output_path: str | None = None
    total_rows_added: int = 0
    sheet_summary: dict[str, int] = {}
    mismatches: list[SheetMismatchDetail] = []
    warnings: list[str] = []
    error: dict | None = None


# ── 辅助函数 ──


def _safe_value(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _find_last_data_row(ws, max_col: int = 0) -> int:
    """从下往上找到最后一个有数据的行号。"""
    col_limit = max_col or ws.max_column
    for row_idx in range(ws.max_row, 0, -1):
        for col_idx in range(1, col_limit + 1):
            if ws.cell(row=row_idx, column=col_idx).value is not None:
                return row_idx
    return 1


def _count_data_rows(ws, header_row: int = 1) -> int:
    """统计 header 之后的非空数据行数。"""
    count = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        has_data = any(ws.cell(row=row_idx, column=c).value is not None for c in range(1, ws.max_column + 1))
        if has_data:
            count += 1
    return count


def _extract_images_from_sheet(ws) -> list[dict]:
    """提取 sheet 中所有图片，返回 [{row, col, data}]。"""
    images = []
    if not hasattr(ws, "_images") or not ws._images:
        return images

    for img in ws._images:
        anchor = img.anchor
        from_anchor = getattr(anchor, "_from", None) or getattr(anchor, "col", None)

        if from_anchor is not None:
            row = getattr(from_anchor, "row", None)
            col = getattr(from_anchor, "col", None)
        else:
            row = None
            col = None

        img_data = _get_image_data(img)
        if img_data and row is not None:
            images.append({"row": row, "col": col or 0, "data": img_data})

    return images


def _get_image_data(img) -> bytes | None:
    """从 openpyxl Image 对象获取图片字节。"""
    if hasattr(img, "_data"):
        return img._data()
    try:
        if hasattr(img, "ref"):
            return bytes(img.ref)
    except Exception:
        pass
    return None


def _build_column_map(base_headers: list[str], src_headers: list[str]) -> dict[int, int | None]:
    """通过表头名称建立 base 列索引 → source 列索引 的映射。"""
    src_name_to_idx = {}
    for i, h in enumerate(src_headers):
        normalized = h.strip().lower() if h else ""
        if normalized:
            src_name_to_idx[normalized] = i

    col_map = {}
    for i, h in enumerate(base_headers):
        normalized = h.strip().lower() if h else ""
        col_map[i] = src_name_to_idx.get(normalized)

    return col_map


def _copy_cell(src_cell, dst_cell):
    """复制单元格的值和样式。"""
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def _add_image_to_sheet(ws, img_data: bytes, row_1based: int, col_0based: int):
    """将图片添加到 sheet 的指定单元格位置。"""
    stream = BytesIO(img_data)
    img = OpenpyxlImage(stream)
    col_letter = get_column_letter(col_0based + 1)
    img.anchor = f"{col_letter}{row_1based}"
    ws.add_image(img)


# ── 端点 ──


@router.post("/merge-excel-info", response_model=MergeInfoResponse)
def merge_excel_info(req: MergeInfoRequest):
    """扫描多个 xlsx 文件的 sheet 结构信息。"""
    try:
        files_info = []
        per_file_sheet_names = []

        for fp in req.file_paths:
            path = Path(fp)
            if not path.exists():
                continue
            wb = load_workbook(fp, data_only=True)
            sheets = []
            sheet_names = set()
            for ws in wb.worksheets:
                headers = [_safe_value(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
                has_imgs = bool(hasattr(ws, "_images") and ws._images)
                sheets.append(SheetInfo(
                    sheet_name=ws.title,
                    header_row=headers,
                    data_row_count=_count_data_rows(ws),
                    has_images=has_imgs,
                    column_count=ws.max_column,
                ))
                sheet_names.add(ws.title)
            wb.close()
            per_file_sheet_names.append(sheet_names)
            files_info.append(FileSheetInfo(
                file_path=str(path),
                file_name=path.name,
                sheets=sheets,
            ))

        # 取所有文件的 sheet 交集
        if per_file_sheet_names and len(per_file_sheet_names) == len(files_info):
            common = set.intersection(*per_file_sheet_names)
        else:
            common = set()

        return MergeInfoResponse(success=True, files=files_info, common_sheets=sorted(common))
    except Exception as e:
        return MergeInfoResponse(success=False, files=[], common_sheets=[], error={"code": "MERGE_INFO_ERROR", "message": str(e)})


@router.post("/merge-excel", response_model=MergeExcelResponse)
def merge_excel(req: MergeExcelRequest):
    """执行 Excel 合并：将 source_files 的数据行追加到 base_file 的对应 sheet。"""
    try:
        base_path = Path(req.base_file)
        output_path = Path(req.output_path)

        if not base_path.exists():
            return MergeExcelResponse(success=False, error={"code": "FILE_NOT_FOUND", "message": f"基准文件不存在: {base_path}"})

        # 复制 base 文件作为输出
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(base_path), str(output_path))

        # 打开副本进行写入
        wb_out = load_workbook(str(output_path))

        mismatches = []
        warnings = []
        total_rows = 0
        sheet_summary = {}

        for sheet_name in req.selected_sheets:
            if sheet_name not in wb_out.sheetnames:
                warnings.append(f"Sheet '{sheet_name}' 不在基准文件中，已跳过")
                continue

            ws_out = wb_out[sheet_name]
            out_last_row = _find_last_data_row(ws_out)
            header_row_idx = 1

            # 读取 base 表头
            base_headers = [_safe_value(ws_out.cell(row=header_row_idx, column=c).value) for c in range(1, ws_out.max_column + 1)]

            rows_added = 0

            for src_path_str in req.source_files:
                src_path = Path(src_path_str)
                if not src_path.exists():
                    warnings.append(f"文件不存在: {src_path.name}")
                    continue

                wb_src = load_workbook(str(src_path), data_only=False)
                if sheet_name not in wb_src.sheetnames:
                    wb_src.close()
                    continue

                ws_src = wb_src[sheet_name]

                # 读取 source 表头
                src_headers = [_safe_value(ws_src.cell(row=header_row_idx, column=c).value) for c in range(1, ws_src.max_column + 1)]

                # 建立列映射
                col_map = _build_column_map(base_headers, src_headers)

                # 检测列不匹配
                base_set = {h for h in base_headers if h}
                src_set = {h for h in src_headers if h}
                missing = sorted(base_set - src_set)
                extra = sorted(src_set - base_set)
                if missing or extra:
                    mismatches.append(SheetMismatchDetail(
                        sheet_name=sheet_name,
                        file_name=src_path.name,
                        base_headers=[h for h in base_headers if h],
                        file_headers=[h for h in src_headers if h],
                        missing_in_file=missing,
                        extra_in_file=extra,
                    ))

                # 提取 source 图片
                src_images = _extract_images_from_sheet(ws_src)

                # 行映射：source row -> output row
                src_row_to_out_row = {}

                # 复制数据行
                for src_row_idx in range(header_row_idx + 1, ws_src.max_row + 1):
                    # 检查是否为空行
                    is_empty = all(
                        ws_src.cell(row=src_row_idx, column=c).value is None
                        for c in range(1, ws_src.max_column + 1)
                    )
                    if is_empty:
                        continue

                    out_last_row += 1
                    rows_added += 1
                    src_row_to_out_row[src_row_idx] = out_last_row

                    # 按列映射复制
                    for base_col_idx in range(len(base_headers)):
                        if not base_headers[base_col_idx]:
                            continue
                        src_col_idx = col_map.get(base_col_idx)
                        if src_col_idx is not None:
                            src_cell = ws_src.cell(row=src_row_idx, column=src_col_idx + 1)
                            dst_cell = ws_out.cell(row=out_last_row, column=base_col_idx + 1)
                            _copy_cell(src_cell, dst_cell)

                # 复制图片（按行映射偏移）
                for img_info in src_images:
                    src_row = img_info["row"] + 1  # openpyxl _from.row 是 0-based
                    out_row = src_row_to_out_row.get(src_row)
                    if out_row is not None:
                        _add_image_to_sheet(ws_out, img_info["data"], out_row, img_info["col"])

                wb_src.close()

            total_rows += rows_added
            sheet_summary[sheet_name] = rows_added

        wb_out.save(str(output_path))
        wb_out.close()

        if total_rows > 0:
            warnings.append("合并完成。如存在公式引用，请打开文件检查并修正公式目标单元格。")

        return MergeExcelResponse(
            success=True,
            output_path=str(output_path),
            total_rows_added=total_rows,
            sheet_summary=sheet_summary,
            mismatches=mismatches,
            warnings=warnings,
        )
    except Exception as e:
        return MergeExcelResponse(success=False, error={"code": "MERGE_ERROR", "message": str(e)})
